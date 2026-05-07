"""
Conversión de formatos a PDF y validación de archivos.
- Validación de PDF, Excel, imágenes
- Conversión Excel (.xlsx/.xls/.xlsm) a PDF
- Conversión de imágenes (jpg/png/tiff/etc.) a PDF
"""
import os
import io
import pandas as pd
from PIL import Image
import fitz
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, PageBreak, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
import re
import logging

logger = logging.getLogger(__name__)


# Tamaño máximo por defecto (MB). Ajustable al integrar.
MAX_FILE_SIZE_MB = 100


# ===========================================================================
# Validadores
# ===========================================================================

def validar_pdf(file_bytes: bytes) -> bool:
    if not file_bytes.startswith(b'%PDF'):
        return False
    try:
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        doc.close()
        return True
    except Exception:
        return False

def validar_excel(file_bytes: bytes, nombre_archivo: str) -> bool:
    extension = os.path.splitext(nombre_archivo.lower())[1]
    if extension in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
        if not file_bytes.startswith(b'PK\x03\x04'):
            return False
    elif extension in ['.xls', '.xlsb']:
        if not file_bytes.startswith(b'\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1'):
            return False
    try:
        engine = 'xlrd' if extension == '.xls' else 'openpyxl'
        pd.read_excel(io.BytesIO(file_bytes), engine=engine, nrows=0)
        return True
    except Exception:
        return False

def validar_tamano_archivo(file_bytes: bytes, max_size_mb: int = MAX_FILE_SIZE_MB) -> None:
    file_size_mb = len(file_bytes) / (1024 * 1024)
    if file_size_mb > max_size_mb:
        raise ValueError(f"Archivo excede el tamaño máximo de {max_size_mb}MB ({file_size_mb:.1f}MB)")

def es_archivo_excel(nombre_archivo: str) -> bool:
    extensiones_excel = ['.xls', '.xlsx', '.xlsm', '.xlsb', '.xltx', '.xltm']
    return os.path.splitext(nombre_archivo.lower())[1] in extensiones_excel

def es_archivo_imagen(nombre_archivo: str) -> bool:
    extensiones_imagen = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.tif', '.webp']
    return os.path.splitext(nombre_archivo.lower())[1] in extensiones_imagen

def validar_imagen(file_bytes: bytes, nombre_archivo: str) -> bool:
    try:
        imagen = Image.open(io.BytesIO(file_bytes))
        imagen.verify()
        return True
    except Exception:
        return False


# ===========================================================================
# Conversión (resto del archivo viene de pdf_converter.py)
# ===========================================================================


# Setup fonts
try:
    pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
    FUENTE_PRINCIPAL = 'STSong-Light'
except Exception:
    FUENTE_PRINCIPAL = 'Helvetica'


def formatear_celda(valor):
    """Formatea el valor de una celda para su representación en el PDF."""
    if pd.isna(valor):
        return ""

    if isinstance(valor, (float, int)):
        try:
            val_float = float(valor)
            if val_float.is_integer():
                return str(int(val_float))
            return str(val_float)
        except Exception:
            return str(valor)

    if isinstance(valor, pd.Timestamp):
        return valor.strftime('%Y-%m-%d')

    return str(valor)


def limpiar_dataframe(df):
    """Limpia el DataFrame eliminando columnas y filas vacías, detectando gaps."""
    df = df.fillna('')
    df = df.astype(str)
    df = df.replace('nan', '')

    densidades = []
    for col in df.columns:
        celdas_con_dato = (df[col].str.strip().str.len() > 0).sum()
        densidades.append(celdas_con_dato)

    umbral = 2

    hay_gap = False
    gap_actual = 0
    for d in densidades:
        if d < umbral:
            gap_actual += 1
            if gap_actual >= 3:
                hay_gap = True
                break
        else:
            gap_actual = 0

    if hay_gap:
        cols_con_datos_idx = [i for i, d in enumerate(densidades) if d >= umbral]

        if not cols_con_datos_idx:
            return pd.DataFrame()

        primer_col = cols_con_datos_idx[0]
        fin_bloque = cols_con_datos_idx[-1]

        for i in range(primer_col, len(densidades)):
            gap = 0
            for j in range(i, min(i + 5, len(densidades))):
                if densidades[j] < umbral:
                    gap += 1
                else:
                    break
            if gap >= 3:
                fin_bloque = i - 1
                break

        cols_seleccionadas = [i for i in range(primer_col, fin_bloque + 1) if densidades[i] >= umbral]
        if not cols_seleccionadas:
            return pd.DataFrame()
        df = df.iloc[:, cols_seleccionadas]
    else:
        cols_con_datos = []
        for col in df.columns:
            if df[col].str.strip().str.len().sum() > 0:
                cols_con_datos.append(col)
        df = df[cols_con_datos]

    df = df[df.apply(lambda row: row.str.strip().str.len().sum() > 0, axis=1)]

    return df.reset_index(drop=True)


def calcular_anchos_columnas_mejorado(df, disponible_width):
    """Calcula anchos de columnas óptimos basados en el contenido."""
    if df.empty:
        return []

    num_cols = len(df.columns)
    max_lens = []

    for col_idx in range(num_cols):
        col_data = df.iloc[:, col_idx].astype(str)
        longitudes = col_data.str.len()
        p90 = longitudes.quantile(0.9) if len(longitudes) > 0 else 0
        max_lens.append(max(p90, 5))

    total_chars = sum(max_lens)
    if total_chars == 0:
        total_chars = 1

    col_widths = [(l / total_chars) * disponible_width for l in max_lens]

    if num_cols <= 6:
        min_width = 0.8 * inch
        max_width = 5.0 * inch
    elif num_cols <= 10:
        min_width = 0.5 * inch
        max_width = 4.0 * inch
    else:
        min_width = 0.4 * inch
        max_width = 3.0 * inch

    col_widths = [max(min(w, max_width), min_width) for w in col_widths]

    current_total = sum(col_widths)
    if current_total > 0:
        factor = disponible_width / current_total
        col_widths = [w * factor for w in col_widths]

    return col_widths


def _obtener_anchos_excel(ws, num_cols):
    """Obtiene los anchos de columna definidos en el Excel como lista proporcional."""
    from openpyxl.utils import get_column_letter
    anchos = []
    for col_idx in range(1, num_cols + 1):
        col_letter = get_column_letter(col_idx)
        w = ws.column_dimensions[col_letter].width
        if w is None or w == 0:
            w = 8.43  # default Excel width
        anchos.append(w)
    return anchos


def _detectar_fila_encabezado_tabla(ws):
    """Detecta la fila de encabezado de la tabla de datos en la hoja.

    Busca la primera fila donde CADA celda tiene su propio valor real (no merged)
    y al menos el 60% de las columnas tienen contenido. Ignora filas donde el
    contenido viene de celdas combinadas que abarcan muchas columnas.
    """
    # Construir un set de celdas que son parte de un merge amplio (>3 cols)
    celdas_merge_amplio = set()
    for mc in ws.merged_cells.ranges:
        span_cols = mc.max_col - mc.min_col + 1
        if span_cols > 3:  # merge que abarca más de 3 columnas = encabezado
            for r in range(mc.min_row, mc.max_row + 1):
                for c in range(mc.min_col, mc.max_col + 1):
                    celdas_merge_amplio.add((r, c))

    num_cols = ws.max_column or 1

    for row_idx in range(1, ws.max_row + 1):
        # Contar cuántas celdas de esta fila están en merges amplios
        en_merge_amplio = sum(1 for c in range(1, num_cols + 1) if (row_idx, c) in celdas_merge_amplio)
        if en_merge_amplio > num_cols * 0.3:
            continue  # fila dominada por merges amplios, es encabezado corporativo

        # Contar celdas con valor real
        celdas_con_valor = 0
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None and str(cell.value).strip():
                celdas_con_valor += 1

        if celdas_con_valor >= num_cols * 0.6:
            return row_idx

    return 1


def _extraer_encabezado_corporativo(ws, fila_inicio_tabla):
    """Extrae el encabezado corporativo (filas antes de la tabla) como líneas de texto."""
    lineas = []
    vistas = set()

    for row_idx in range(1, fila_inicio_tabla):
        textos_fila = []
        for col_idx in range(1, (ws.max_column or 1) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            if val is not None:
                s = str(val).strip()
                if s and s not in vistas:
                    vistas.add(s)
                    textos_fila.append(s)

        if textos_fila:
            linea = '    '.join(textos_fila)
            lineas.append(linea)

    return lineas


def _convertir_excel_a_pdf_avanzado(excel_bytes: bytes, nombre_archivo: str) -> bytes:
    """Conversión avanzada Excel a PDF usando openpyxl para preservar estructura."""
    from openpyxl import load_workbook
    from reportlab.platypus import Spacer

    wb = load_workbook(io.BytesIO(excel_bytes), data_only=True)

    buffer = io.BytesIO()
    styles = getSampleStyleSheet()

    estilo_encabezado = ParagraphStyle(
        'EncabezadoDoc',
        fontName=FUENTE_PRINCIPAL,
        fontSize=7,
        leading=9,
        wordWrap='CJK',
    )

    estilo_titulo_hoja = ParagraphStyle(
        'TituloHojaAvz',
        parent=styles['Heading1'],
        fontName=FUENTE_PRINCIPAL,
        fontSize=14,
        spaceAfter=8,
    )

    font_sizes = [7, 6, 5, 4.5]
    last_error = None

    for font_size in font_sizes:
        try:
            elements = []
            leading = font_size + 2

            estilo_celda = ParagraphStyle(
                f'CeldaAvz_{font_size}',
                fontName=FUENTE_PRINCIPAL,
                fontSize=font_size,
                leading=leading,
                wordWrap='CJK',
                splitLongWords=1,
            )

            estilo_celda_header = ParagraphStyle(
                f'CeldaAvzH_{font_size}',
                fontName=FUENTE_PRINCIPAL,
                fontSize=font_size + 0.5,
                leading=leading + 1,
                wordWrap='CJK',
                splitLongWords=1,
                textColor=colors.black,
            )

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if ws.max_row is None or ws.max_row < 2:
                    continue

                num_cols = ws.max_column or 14

                elements.append(Paragraph(f"Hoja: {sheet_name}", estilo_titulo_hoja))

                # Detectar dónde empieza la tabla de datos (fila Excel, 1-based)
                fila_header = _detectar_fila_encabezado_tabla(ws)

                # Renderizar encabezado corporativo como texto simple
                if fila_header > 1:
                    lineas_enc = _extraer_encabezado_corporativo(ws, fila_header)
                    for linea in lineas_enc:
                        texto_html = linea.replace('\n', '<br/>').replace('&', '&amp;')
                        elements.append(Paragraph(texto_html, estilo_encabezado))
                    elements.append(Spacer(1, 6))

                # Leer datos de tabla desde la fila de encabezado (sin propagar merges amplios)
                celdas_merge_amplio = set()
                for mc in ws.merged_cells.ranges:
                    if (mc.max_col - mc.min_col + 1) > 3:
                        for r in range(mc.min_row, mc.max_row + 1):
                            for c in range(mc.min_col, mc.max_col + 1):
                                celdas_merge_amplio.add((r, c))

                # Resolver merges pequeños (<=3 cols) para la zona de datos
                merge_pequeño = {}
                for mc in ws.merged_cells.ranges:
                    if (mc.max_col - mc.min_col + 1) <= 3:
                        val = ws.cell(mc.min_row, mc.min_col).value
                        for r in range(mc.min_row, mc.max_row + 1):
                            for c in range(mc.min_col, mc.max_col + 1):
                                if r != mc.min_row or c != mc.min_col:
                                    merge_pequeño[(r, c)] = val

                table_data_raw = []
                for row_idx in range(fila_header, ws.max_row + 1):
                    fila = []
                    for col_idx in range(1, num_cols + 1):
                        if (row_idx, col_idx) in celdas_merge_amplio:
                            fila.append(None)
                        elif (row_idx, col_idx) in merge_pequeño:
                            fila.append(merge_pequeño[(row_idx, col_idx)])
                        else:
                            fila.append(ws.cell(row=row_idx, column=col_idx).value)
                    table_data_raw.append(fila)

                if not table_data_raw:
                    continue

                # Determinar columnas con datos en la zona de tabla
                col_density = [0] * num_cols
                for fila in table_data_raw:
                    for ci in range(num_cols):
                        if ci < len(fila) and fila[ci] is not None and str(fila[ci]).strip():
                            col_density[ci] += 1

                cols_activas = [i for i in range(num_cols) if col_density[i] > 0]
                if not cols_activas:
                    continue

                # Construir datos de tabla
                table_data = []
                for row_idx, fila in enumerate(table_data_raw):
                    fila_filtrada = []
                    tiene_dato = False
                    for ci in cols_activas:
                        val = fila[ci] if ci < len(fila) else None
                        texto = formatear_celda(val) if val is not None else ""
                        texto = texto.replace('\n', '<br/>')
                        if texto.strip():
                            tiene_dato = True
                        if row_idx == 0:
                            fila_filtrada.append(Paragraph(texto, estilo_celda_header))
                        else:
                            fila_filtrada.append(Paragraph(texto, estilo_celda))
                    if tiene_dato:
                        table_data.append(fila_filtrada)

                if not table_data:
                    continue

                # Anchos de columna basados en Excel
                anchos_excel = _obtener_anchos_excel(ws, num_cols)
                anchos_activos = [anchos_excel[i] for i in cols_activas]
                total_excel = sum(anchos_activos)

                ancho_util = landscape(A4)[0] - 0.6 * inch
                if total_excel > 0:
                    col_widths = [(a / total_excel) * ancho_util for a in anchos_activos]
                else:
                    col_widths = [ancho_util / len(cols_activas)] * len(cols_activas)

                min_w = 0.3 * inch
                col_widths = [max(w, min_w) for w in col_widths]
                factor = ancho_util / sum(col_widths)
                col_widths = [w * factor for w in col_widths]

                table = Table(table_data, colWidths=col_widths, repeatRows=1)

                estilo_grid = [
                    ('FONTNAME', (0, 0), (-1, -1), FUENTE_PRINCIPAL),
                    ('FONTSIZE', (0, 0), (-1, -1), font_size),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.90, 0.90, 0.90)),
                    ('LEFTPADDING', (0, 0), (-1, -1), 2),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                    ('TOPPADDING', (0, 0), (-1, -1), 1),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                ]

                # Alinear columnas numéricas a la derecha (últimas 4)
                if len(cols_activas) >= 4:
                    for ci in range(len(cols_activas) - 4, len(cols_activas)):
                        estilo_grid.append(('ALIGN', (ci, 1), (ci, -1), 'RIGHT'))

                table.setStyle(TableStyle(estilo_grid))
                elements.append(table)
                elements.append(PageBreak())

            if not elements:
                return None

            buffer.seek(0)
            buffer.truncate()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=landscape(A4),
                leftMargin=0.3 * inch,
                rightMargin=0.3 * inch,
                topMargin=0.3 * inch,
                bottomMargin=0.3 * inch,
            )
            doc.build(elements)

            if font_size != font_sizes[0]:
                logger.info(f"PDF avanzado generado con fuente {font_size}pt para {nombre_archivo}")

            return buffer.getvalue()

        except Exception as build_err:
            err_msg = str(build_err)
            if 'too large on page' in err_msg or 'too large for frame' in err_msg:
                last_error = build_err
                logger.warning(
                    f"Tabla demasiado grande con fuente {font_size}pt (avanzado) para {nombre_archivo}, "
                    f"reintentando..."
                )
                continue
            else:
                raise

    if last_error:
        raise last_error
    return None


def _detectar_fila_encabezado_tabla_xls(ws, merged_cells):
    """Detecta la fila de encabezado de la tabla de datos en una hoja xlrd.

    Retorna el indice de fila (0-based) o -1 si la hoja no tiene estructura tabular.
    Valida que despues de la fila candidata haya suficientes filas con datos
    para confirmar que es una tabla real.
    """
    num_cols = ws.ncols or 1

    celdas_merge_amplio = set()
    for rlo, rhi, clo, chi in merged_cells:
        span_cols = chi - clo
        if span_cols > 3:
            for r in range(rlo, rhi):
                for c in range(clo, chi):
                    celdas_merge_amplio.add((r, c))

    # Contar cuantas columnas tienen datos en al menos alguna fila
    col_tiene_dato = [False] * num_cols
    for row_idx in range(ws.nrows):
        for col_idx in range(num_cols):
            if (row_idx, col_idx) not in celdas_merge_amplio:
                val = ws.cell_value(row_idx, col_idx)
                if val is not None and str(val).strip():
                    col_tiene_dato[col_idx] = True
    cols_con_datos = sum(col_tiene_dato)

    umbral = max(int(cols_con_datos * 0.4), 4)

    for row_idx in range(ws.nrows):
        en_merge_amplio = sum(1 for c in range(num_cols) if (row_idx, c) in celdas_merge_amplio)
        if en_merge_amplio > num_cols * 0.3:
            continue

        celdas_con_valor = 0
        for col_idx in range(num_cols):
            val = ws.cell_value(row_idx, col_idx)
            if val is not None and str(val).strip():
                celdas_con_valor += 1

        if celdas_con_valor >= umbral:
            # Validar: debe haber al menos 5 filas con datos despues de esta
            filas_con_datos = 0
            for check_row in range(row_idx + 1, min(row_idx + 20, ws.nrows)):
                check_vals = 0
                for col_idx in range(num_cols):
                    val = ws.cell_value(check_row, col_idx)
                    if val is not None and str(val).strip():
                        check_vals += 1
                if check_vals >= 3:
                    filas_con_datos += 1

            if filas_con_datos >= 5:
                return row_idx

    return -1


def _extraer_encabezado_corporativo_xls(ws, fila_inicio_tabla):
    """Extrae el encabezado corporativo de una hoja xlrd."""
    lineas = []
    vistas = set()

    for row_idx in range(fila_inicio_tabla):
        textos_fila = []
        for col_idx in range(ws.ncols):
            val = ws.cell_value(row_idx, col_idx)
            if val is not None:
                s = str(val).strip()
                if s and s not in vistas:
                    vistas.add(s)
                    textos_fila.append(s)
        if textos_fila:
            linea = '    '.join(textos_fila)
            lineas.append(linea)

    return lineas


def _renderizar_hoja_layout_xls(ws, merged_cells):
    """Extrae el contenido de una hoja de layout (sin estructura tabular) como lineas de texto.

    Resuelve celdas combinadas y agrupa por fila, eliminando duplicados.
    """
    # Construir mapa de merge: para cada celda que es parte de un merge,
    # apuntar al valor de la celda origen
    merge_map = {}
    for rlo, rhi, clo, chi in merged_cells:
        val = ws.cell_value(rlo, clo)
        for r in range(rlo, rhi):
            for c in range(clo, chi):
                merge_map[(r, c)] = val

    lineas = []
    for row_idx in range(ws.nrows):
        textos_fila = []
        vistas_fila = set()
        for col_idx in range(ws.ncols):
            if (row_idx, col_idx) in merge_map:
                val = merge_map[(row_idx, col_idx)]
            else:
                val = ws.cell_value(row_idx, col_idx)

            if val is not None:
                s = str(val).strip().replace('\n', ' ')
                if s and s not in vistas_fila:
                    vistas_fila.add(s)
                    textos_fila.append(s)

        if textos_fila:
            lineas.append('    '.join(textos_fila))

    return lineas


def _obtener_anchos_xls(ws, num_cols):
    """Obtiene los anchos de columna definidos en un archivo .xls via xlrd."""
    anchos = []
    for col_idx in range(num_cols):
        try:
            w = ws.colinfo_map.get(col_idx)
            if w is not None:
                anchos.append(w.width / 256.0)
            else:
                anchos.append(8.43)
        except Exception:
            anchos.append(8.43)
    return anchos


def _convertir_xls_a_pdf_avanzado(excel_bytes: bytes, nombre_archivo: str) -> bytes:
    """Conversión avanzada de .xls a PDF usando xlrd para preservar estructura."""
    import xlrd
    from reportlab.platypus import Spacer

    wb = xlrd.open_workbook(file_contents=excel_bytes, formatting_info=True)

    buffer = io.BytesIO()
    styles = getSampleStyleSheet()

    estilo_encabezado = ParagraphStyle(
        'EncabezadoDocXls',
        fontName=FUENTE_PRINCIPAL,
        fontSize=7,
        leading=9,
        wordWrap='CJK',
    )

    estilo_titulo_hoja = ParagraphStyle(
        'TituloHojaXls',
        parent=styles['Heading1'],
        fontName=FUENTE_PRINCIPAL,
        fontSize=14,
        spaceAfter=8,
    )

    font_sizes = [7, 6, 5, 4.5]
    last_error = None

    for font_size in font_sizes:
        try:
            elements = []
            leading = font_size + 2

            estilo_celda = ParagraphStyle(
                f'CeldaXls_{font_size}',
                fontName=FUENTE_PRINCIPAL,
                fontSize=font_size,
                leading=leading,
                wordWrap='CJK',
                splitLongWords=1,
            )

            estilo_celda_header = ParagraphStyle(
                f'CeldaXlsH_{font_size}',
                fontName=FUENTE_PRINCIPAL,
                fontSize=font_size + 0.5,
                leading=leading + 1,
                wordWrap='CJK',
                splitLongWords=1,
                textColor=colors.black,
            )

            for sheet_idx in range(wb.nsheets):
                ws = wb.sheet_by_index(sheet_idx)
                sheet_name = ws.name

                if ws.nrows < 2:
                    continue

                num_cols = ws.ncols or 14
                merged_cells = ws.merged_cells

                elements.append(Paragraph(f"Hoja: {sheet_name}", estilo_titulo_hoja))

                fila_header = _detectar_fila_encabezado_tabla_xls(ws, merged_cells)

                # Si no se detectó estructura tabular, renderizar como texto
                if fila_header == -1:
                    lineas = _renderizar_hoja_layout_xls(ws, merged_cells)
                    for linea in lineas:
                        texto_html = linea.replace('\n', '<br/>').replace('&', '&amp;')
                        elements.append(Paragraph(texto_html, estilo_encabezado))
                    elements.append(PageBreak())
                    continue

                # Construir sets de celdas en merges amplios y pequeños
                celdas_merge_amplio = set()
                celdas_merge_secundaria = set()
                for rlo, rhi, clo, chi in merged_cells:
                    span_cols = chi - clo
                    if span_cols > 3:
                        for r in range(rlo, rhi):
                            for c in range(clo, chi):
                                celdas_merge_amplio.add((r, c))
                    else:
                        for r in range(rlo, rhi):
                            for c in range(clo, chi):
                                if r != rlo or c != clo:
                                    celdas_merge_secundaria.add((r, c))

                # Leer datos de tabla
                table_data_raw = []
                for row_idx in range(fila_header, ws.nrows):
                    fila = []
                    for col_idx in range(num_cols):
                        if (row_idx, col_idx) in celdas_merge_amplio:
                            fila.append(None)
                        elif (row_idx, col_idx) in celdas_merge_secundaria:
                            fila.append(None)
                        else:
                            fila.append(ws.cell_value(row_idx, col_idx))
                    table_data_raw.append(fila)

                if not table_data_raw:
                    continue

                # Determinar columnas con datos suficientes
                col_density = [0] * num_cols
                for fila in table_data_raw:
                    for ci in range(num_cols):
                        if ci < len(fila) and fila[ci] is not None and str(fila[ci]).strip():
                            col_density[ci] += 1

                total_data_rows = len(table_data_raw)
                min_density = max(int(total_data_rows * 0.05), 3)
                cols_activas = [i for i in range(num_cols) if col_density[i] >= min_density]
                if not cols_activas:
                    cols_activas = [i for i in range(num_cols) if col_density[i] > 0]
                if not cols_activas:
                    continue

                # Filtrar columnas triviales (contenido es solo puntuacion o 1 caracter)
                cols_filtradas = []
                for ci in cols_activas:
                    max_len_contenido = 0
                    for fila in table_data_raw:
                        val = fila[ci] if ci < len(fila) else None
                        if val is not None:
                            txt = re.sub(r'\s+', ' ', str(val)).strip()
                            if len(txt) > max_len_contenido:
                                max_len_contenido = len(txt)
                    if max_len_contenido >= 2:
                        cols_filtradas.append(ci)
                if cols_filtradas:
                    cols_activas = cols_filtradas

                # --- Buscar encabezados de columna en zona pre-datos ---
                merge_origin_map = {}
                for rlo, rhi, clo, chi in merged_cells:
                    for r in range(rlo, rhi):
                        for c in range(clo, chi):
                            if r != rlo or c != clo:
                                merge_origin_map[(r, c)] = (rlo, clo)

                min_ac = min(cols_activas)
                max_ac = max(cols_activas)
                rango_ac = max(max_ac - min_ac, 1)

                # Columnas de alta densidad (>20% de las filas tienen datos)
                cols_alta_densidad = set()
                for ci in cols_activas:
                    if col_density[ci] > total_data_rows * 0.2:
                        cols_alta_densidad.add(ci)

                candidatas_enc = []
                for row_idx in range(max(0, fila_header - 5), fila_header):
                    vals_por_col = {}
                    origenes_vistos = set()
                    for col_idx in range(num_cols):
                        if (row_idx, col_idx) in merge_origin_map:
                            orig = merge_origin_map[(row_idx, col_idx)]
                            if orig[0] != row_idx:
                                continue
                            if orig in origenes_vistos:
                                continue
                            origenes_vistos.add(orig)
                            val = ws.cell_value(orig[0], orig[1])
                            src_col = orig[1]
                        else:
                            val = ws.cell_value(row_idx, col_idx)
                            src_col = col_idx

                        if val is None or not str(val).strip():
                            continue
                        text = str(val).strip().replace('\n', ' ')
                        if len(text) > 25 or len(text) < 3:
                            continue
                        if text.replace('.', '').replace(',', '').replace(' ', '').isdigit():
                            continue

                        mejor_ac = None
                        min_dist = float('inf')
                        for ac in cols_activas:
                            d = abs(src_col - ac)
                            if d < min_dist:
                                min_dist = d
                                mejor_ac = ac
                        if mejor_ac is not None and min_dist <= 3:
                            vals_por_col[mejor_ac] = text

                    if len(vals_por_col) >= 3:
                        val_cols = sorted(vals_por_col.keys())
                        spread = val_cols[-1] - val_cols[0]
                        # Contar cuantos valores caen en columnas de alta densidad
                        en_alta_densidad = sum(1 for c in vals_por_col if c in cols_alta_densidad)
                        candidatas_enc.append({
                            'row': row_idx,
                            'vals': vals_por_col,
                            'spread': spread,
                            'en_alta_densidad': en_alta_densidad,
                            'n_cols': len(vals_por_col),
                        })

                # Seleccionar las mejores filas de encabezado:
                # Priorizar fila con mayor cobertura de columnas de alta densidad
                filas_encabezado_cols = []
                primera_fila_enc_col = fila_header

                if candidatas_enc:
                    # Ordenar por cobertura de cols alta densidad, luego por n_cols
                    candidatas_enc.sort(key=lambda x: (x['en_alta_densidad'], x['n_cols']), reverse=True)
                    mejor = candidatas_enc[0]

                    # Solo aceptar si cubre al menos 40% de las cols de alta densidad
                    if len(cols_alta_densidad) > 0:
                        cobertura = mejor['en_alta_densidad'] / len(cols_alta_densidad)
                    else:
                        cobertura = 1.0

                    if cobertura >= 0.35 and mejor['spread'] >= rango_ac * 0.3:
                        filas_encabezado_cols.append(mejor['vals'])
                        primera_fila_enc_col = mejor['row']

                        # Incluir filas adyacentes que tambien califiquen
                        for cand in candidatas_enc[1:]:
                            if abs(cand['row'] - mejor['row']) <= 1 and cand['en_alta_densidad'] >= 2:
                                filas_encabezado_cols.append(cand['vals'])
                                if cand['row'] < primera_fila_enc_col:
                                    primera_fila_enc_col = cand['row']

                        # Ordenar por numero de fila
                        if len(filas_encabezado_cols) > 1:
                            temp = [(f, r) for f, r in zip(filas_encabezado_cols,
                                    [c['row'] for c in candidatas_enc if c['vals'] in filas_encabezado_cols])]
                            temp.sort(key=lambda x: x[1])
                            filas_encabezado_cols = [t[0] for t in temp]

                # --- Encabezado corporativo: solo hasta donde empiezan los enc. de columna ---
                enc_limite = primera_fila_enc_col if filas_encabezado_cols else fila_header
                if enc_limite > 0:
                    lineas_enc = _extraer_encabezado_corporativo_xls(ws, enc_limite)
                    for linea in lineas_enc:
                        texto_html = linea.replace('\n', '<br/>').replace('&', '&amp;')
                        elements.append(Paragraph(texto_html, estilo_encabezado))
                    elements.append(Spacer(1, 6))

                # Construir filas de encabezado como Paragraphs
                header_table_data = []
                for hrow in filas_encabezado_cols:
                    fila = []
                    for ci in cols_activas:
                        texto = hrow.get(ci, "")
                        fila.append(Paragraph(texto, estilo_celda_header))
                    header_table_data.append(fila)
                num_header_rows = len(header_table_data)

                # Construir datos de tabla
                table_data = []
                for row_idx, fila in enumerate(table_data_raw):
                    fila_filtrada = []
                    tiene_dato = False
                    for ci in cols_activas:
                        val = fila[ci] if ci < len(fila) else None
                        texto = formatear_celda(val) if val is not None else ""
                        texto = re.sub(r'\s+', ' ', texto).strip()
                        if texto:
                            tiene_dato = True
                        fila_filtrada.append(Paragraph(texto, estilo_celda))
                    if tiene_dato:
                        table_data.append(fila_filtrada)

                if not table_data:
                    continue

                table_data = header_table_data + table_data

                # --- Calculo de anchos de columna ---
                # Recopilar longitud maxima de texto por columna (encabezados + datos)
                max_text_len = [0] * len(cols_activas)
                avg_text_len = [0.0] * len(cols_activas)

                # Encabezados
                for hrow in filas_encabezado_cols:
                    for idx, ci in enumerate(cols_activas):
                        txt = hrow.get(ci, "")
                        if len(txt) > max_text_len[idx]:
                            max_text_len[idx] = len(txt)

                # Datos (muestreo de las primeras 50 filas + ultimas 10)
                sample_rows = table_data_raw[:50] + table_data_raw[-10:]
                for fila in sample_rows:
                    for idx, ci in enumerate(cols_activas):
                        val = fila[ci] if ci < len(fila) else None
                        if val is not None:
                            txt = re.sub(r'\s+', ' ', formatear_celda(val)).strip()
                            if len(txt) > max_text_len[idx]:
                                max_text_len[idx] = len(txt)

                for fila in table_data_raw:
                    for idx, ci in enumerate(cols_activas):
                        val = fila[ci] if ci < len(fila) else None
                        if val is not None:
                            txt = re.sub(r'\s+', ' ', formatear_celda(val)).strip()
                            avg_text_len[idx] += len(txt)
                if total_data_rows > 0:
                    avg_text_len = [t / total_data_rows for t in avg_text_len]

                # Combinar anchos Excel con longitudes de contenido
                anchos_excel = _obtener_anchos_xls(ws, num_cols)
                anchos_activos = [anchos_excel[i] for i in cols_activas]

                # Calcular ancho basado en contenido (caracteres -> unidades Excel aprox)
                CHAR_WIDTH_FACTOR = 1.2
                for idx in range(len(cols_activas)):
                    # Ancho minimo para que quepa el encabezado sin fragmentar
                    ancho_header = max_text_len[idx] * CHAR_WIDTH_FACTOR
                    # Ancho basado en contenido promedio
                    ancho_contenido = max(avg_text_len[idx], max_text_len[idx] * 0.6) * CHAR_WIDTH_FACTOR

                    # Los encabezados deben caber en una linea
                    anchos_activos[idx] = max(anchos_activos[idx], ancho_header, ancho_contenido * 0.4)

                    # Columnas con alta densidad y ancho muy pequeno
                    ci = cols_activas[idx]
                    if col_density[ci] > total_data_rows * 0.5 and anchos_activos[idx] < 3.0:
                        anchos_activos[idx] = max(anchos_activos[idx], 6.0)

                total_ancho = sum(anchos_activos)
                ancho_util = landscape(A4)[0] - 0.6 * inch
                if total_ancho > 0:
                    col_widths = [(a / total_ancho) * ancho_util for a in anchos_activos]
                else:
                    col_widths = [ancho_util / len(cols_activas)] * len(cols_activas)

                min_w = 0.3 * inch
                col_widths = [max(w, min_w) for w in col_widths]
                factor = ancho_util / sum(col_widths)
                col_widths = [w * factor for w in col_widths]

                repeat_rows = num_header_rows if num_header_rows > 0 else 0
                table = Table(table_data, colWidths=col_widths, repeatRows=repeat_rows)

                estilo_grid = [
                    ('FONTNAME', (0, 0), (-1, -1), FUENTE_PRINCIPAL),
                    ('FONTSIZE', (0, 0), (-1, -1), font_size),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
                    ('LEFTPADDING', (0, 0), (-1, -1), 2),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 2),
                    ('TOPPADDING', (0, 0), (-1, -1), 1),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                ]

                if num_header_rows > 0:
                    estilo_grid.append(('BACKGROUND', (0, 0), (-1, num_header_rows - 1),
                                        colors.Color(0.85, 0.85, 0.85)))
                    estilo_grid.append(('FONTSIZE', (0, 0), (-1, num_header_rows - 1),
                                        font_size + 0.5))

                if len(cols_activas) >= 4:
                    first_data_row = num_header_rows
                    for ci in range(len(cols_activas) - 4, len(cols_activas)):
                        estilo_grid.append(('ALIGN', (ci, first_data_row), (ci, -1), 'RIGHT'))

                table.setStyle(TableStyle(estilo_grid))
                elements.append(table)
                elements.append(PageBreak())

            if not elements:
                return None

            buffer.seek(0)
            buffer.truncate()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=landscape(A4),
                leftMargin=0.3 * inch,
                rightMargin=0.3 * inch,
                topMargin=0.3 * inch,
                bottomMargin=0.3 * inch,
            )
            doc.build(elements)

            if font_size != font_sizes[0]:
                logger.info(f"PDF avanzado XLS generado con fuente {font_size}pt para {nombre_archivo}")

            return buffer.getvalue()

        except Exception as build_err:
            err_msg = str(build_err)
            if 'too large on page' in err_msg or 'too large for frame' in err_msg:
                last_error = build_err
                logger.warning(
                    f"Tabla demasiado grande con fuente {font_size}pt (xls avanzado) para {nombre_archivo}, "
                    f"reintentando..."
                )
                continue
            else:
                raise

    if last_error:
        raise last_error
    return None


def convertir_imagen_a_pdf_sync(imagen_bytes: bytes, nombre_archivo: str) -> bytes:
    """Convierte una imagen a PDF de forma síncrona."""
    try:
        imagen = Image.open(io.BytesIO(imagen_bytes))

        if imagen.mode == 'RGBA':
            fondo = Image.new('RGB', imagen.size, (255, 255, 255))
            fondo.paste(imagen, mask=imagen.split()[3])
            imagen = fondo
        elif imagen.mode != 'RGB':
            imagen = imagen.convert('RGB')

        ancho_img, alto_img = imagen.size
        ancho_pagina, alto_pagina = A4
        margen = 30

        ancho_disponible = ancho_pagina - (2 * margen)
        alto_disponible = alto_pagina - (2 * margen)

        escala = min(ancho_disponible / ancho_img, alto_disponible / alto_img)

        ancho_final = ancho_img * escala
        alto_final = alto_img * escala

        x = (ancho_pagina - ancho_final) / 2
        y = (alto_pagina - alto_final) / 2

        buffer = io.BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)

        img_reader = ImageReader(imagen)
        c.drawImage(img_reader, x, y, width=ancho_final, height=alto_final, preserveAspectRatio=True)
        c.save()

        return buffer.getvalue()

    except Exception as e:
        raise ValueError(f"Error al convertir imagen a PDF: {str(e)}")


def convertir_excel_a_pdf_sync(excel_bytes: bytes, nombre_archivo: str) -> bytes:
    """Función síncrona interna para conversión Excel a PDF con soporte avanzado."""
    try:
        extension = os.path.splitext(nombre_archivo.lower())[1]

        # Intentar primero el método avanzado para .xlsx/.xlsm (preserva merged cells y anchos)
        if extension in ['.xlsx', '.xlsm']:
            try:
                resultado = _convertir_excel_a_pdf_avanzado(excel_bytes, nombre_archivo)
                if resultado:
                    logger.info(f"Conversión avanzada exitosa para {nombre_archivo}")
                    return resultado
            except Exception as e_avz:
                logger.warning(
                    f"Conversión avanzada falló para {nombre_archivo}: {str(e_avz)[:150]}. "
                    f"Usando método estándar."
                )

        # Intentar método avanzado para .xls (preserva merged cells y anchos via xlrd)
        if extension == '.xls':
            try:
                resultado = _convertir_xls_a_pdf_avanzado(excel_bytes, nombre_archivo)
                if resultado:
                    logger.info(f"Conversión avanzada XLS exitosa para {nombre_archivo}")
                    return resultado
            except Exception as e_avz:
                logger.warning(
                    f"Conversión avanzada XLS falló para {nombre_archivo}: {str(e_avz)[:150]}. "
                    f"Usando método estándar."
                )


        # Método estándar (fallback o para .xls)
        df_dict = None
        errores = []

        if extension in ['.xlsx', '.xlsm']:
            try:
                df_dict = pd.read_excel(io.BytesIO(excel_bytes), sheet_name=None, engine='openpyxl', header=None)
            except Exception as e1:
                errores.append(f"openpyxl: {str(e1)[:100]}")
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(io.BytesIO(excel_bytes), data_only=True)
                    df_dict = {}
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        data = []
                        for row in ws.iter_rows(values_only=True):
                            data.append(list(row))
                        if data:
                            df_dict[sheet_name] = pd.DataFrame(data)
                except Exception as e2:
                    errores.append(f"openpyxl manual: {str(e2)[:100]}")
        elif extension == '.xls':
            try:
                df_dict = pd.read_excel(io.BytesIO(excel_bytes), sheet_name=None, engine='xlrd', header=None)
            except Exception as e:
                errores.append(f"xlrd: {str(e)[:100]}")
        else:
            try:
                df_dict = pd.read_excel(io.BytesIO(excel_bytes), sheet_name=None, engine='openpyxl', header=None)
            except Exception as e:
                errores.append(f"openpyxl: {str(e)[:100]}")

        if not df_dict:
            logger.warning(f"No se pudo leer Excel {nombre_archivo}: {'; '.join(errores)}")
            raise ValueError(f"Error al leer Excel: {'; '.join(errores)}")

        buffer = io.BytesIO()
        styles = getSampleStyleSheet()

        estilo_titulo = ParagraphStyle(
            'TituloHoja',
            parent=styles['Heading1'],
            fontName=FUENTE_PRINCIPAL,
            fontSize=14,
            spaceAfter=12
        )

        estilo_celda_base = {
            'fontName': FUENTE_PRINCIPAL,
            'fontSize': 9,
            'leading': 11,
            'wordWrap': 'CJK',
            'splitLongWords': 1,
        }

        # Intentar construir el PDF, reduciendo fuente si la tabla es demasiado grande
        font_sizes = [9, 7, 6, 5]
        last_error = None

        for font_size in font_sizes:
            try:
                elements = []
                leading = font_size + 2

                estilo_celda = ParagraphStyle(
                    f'CeldaTabla_{font_size}',
                    **{**estilo_celda_base, 'fontSize': font_size, 'leading': leading}
                )

                for sheet_name, df in df_dict.items():
                    elements.append(Paragraph(f"Hoja: {sheet_name}", estilo_titulo))

                    df_clean = limpiar_dataframe(df)

                    if df_clean.empty:
                        continue

                    data = []
                    for row in df_clean.values.tolist():
                        fila_procesada = []
                        for c in row:
                            texto = formatear_celda(c)
                            fila_procesada.append(Paragraph(texto, estilo_celda))
                        data.append(fila_procesada)

                    if not data:
                        continue

                    ancho_util = landscape(A4)[0] - 0.6*inch
                    col_widths = calcular_anchos_columnas_mejorado(df_clean, ancho_util)

                    table = Table(data, colWidths=col_widths, repeatRows=1)

                    estilo_grid = [
                        ('FONTNAME', (0, 0), (-1, -1), FUENTE_PRINCIPAL),
                        ('FONTSIZE', (0, 0), (-1, -1), font_size),
                        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.93, 0.93, 0.93)),
                        ('LEFTPADDING', (0, 0), (-1, -1), 3),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                        ('TOPPADDING', (0, 0), (-1, -1), 2),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
                    ]

                    table.setStyle(TableStyle(estilo_grid))

                    elements.append(table)
                    elements.append(PageBreak())

                buffer.seek(0)
                buffer.truncate()
                doc = SimpleDocTemplate(
                    buffer,
                    pagesize=landscape(A4),
                    leftMargin=0.3*inch,
                    rightMargin=0.3*inch,
                    topMargin=0.4*inch,
                    bottomMargin=0.4*inch
                )
                doc.build(elements)
                if font_size != 9:
                    logger.info(f"PDF generado con fuente reducida ({font_size}pt) para {nombre_archivo}")
                return buffer.getvalue()

            except Exception as build_err:
                err_msg = str(build_err)
                if 'too large on page' in err_msg or 'too large for frame' in err_msg:
                    last_error = build_err
                    logger.warning(
                        f"Tabla demasiado grande con fuente {font_size}pt para {nombre_archivo}, "
                        f"reintentando con fuente más pequeña..."
                    )
                    continue
                else:
                    raise

        # Si ningún tamaño de fuente funcionó
        raise last_error

    except ValueError:
        raise
    except Exception as e:
        raise ValueError(f"Error al convertir Excel a PDF: {str(e)}")

